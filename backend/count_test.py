import openpyxl

wb_old = openpyxl.load_workbook('Documents/BSCpE_ILO_Skillset_Alignment (1).xlsx', data_only=True)
ws_old = wb_old["ILO–Skillset Alignment"]

subjects_old = set()
for row in ws_old.iter_rows(values_only=True):
    r0 = str(row[0]).strip() if row[0] else ''
    is_subject = (
        row[1] is None and row[2] is None and row[0]
        and 'ILO' not in r0
        and 'BATANGAS' not in r0
        and 'Bachelor' not in r0
        and 'LEGEND' not in r0
        and 'Each ILO' not in r0
        and not r0.startswith('SO')
        and len(r0) > 3
        and '·' in r0
    )
    if is_subject:
        subjects_old.add(r0)

print(f"Old curriculum subjects: {len(subjects_old)}")

wb_new = openpyxl.load_workbook('Documents/BSCpE_ILO_SO_Fully_Distributed_Final.xlsx', data_only=True)
ws_new = wb_new["BSCpE_ILO_SO_Fully_Distributed_"]

subjects_new = set()
for row in ws_new.iter_rows(values_only=True):
    r0 = str(row[0]).strip() if row[0] else ''
    if r0.startswith("Course:"):
        subjects_new.add(r0)
        
print(f"New curriculum subjects: {len(subjects_new)}")
