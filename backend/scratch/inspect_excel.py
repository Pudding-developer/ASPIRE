import openpyxl
from pathlib import Path

excel_path = Path("Documents/BSCpE_ILO_Skillset_Alignment.xlsx")
if not excel_path.exists():
    print(f"File not found: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path)
ws = wb["ILO–Skillset Alignment"]

print(f"Sheet: {ws.title}")
print("-" * 40)

# Look for General Chemistry (SCl 401)
found = False
for row in ws.iter_rows(values_only=True):
    r0 = str(row[0]) if row[0] else ""
    if "General Chemistry" in r0 or "SCl 401" in r0:
        found = True
        print(f"Found Subject: {r0}")
        # Print next 5 rows to see ILOs and columns 5, 6
        start_row = ws.max_row # Not useful here
        
# Actually let's just print a chunk of rows around where General Chemistry is
for i, row in enumerate(ws.iter_rows(values_only=True)):
    r0 = str(row[0]) if row[0] else ""
    if "SCl 401" in r0:
        print(f"Row {i+1}: {row}")
        for j in range(1, 10):
            next_row = list(ws.iter_rows(min_row=i+1+j, max_row=i+1+j, values_only=True))[0]
            print(f"Row {i+1+j}: {next_row}")
        break
