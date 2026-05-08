import asyncio
import json
import re
from pathlib import Path
import openpyxl
from sqlmodel import delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.database import async_session_factory
from app.models.knowledge import KnowledgeChunk
import time

DOCUMENTS_DIR = Path(__file__).parent.parent / "Documents"
ML_DIR = Path(__file__).parent.parent / "ml" / "artifacts"
EXCEL_PATH = DOCUMENTS_DIR / "BSCpE_ILO_Skillset_Alignment.xlsx"

from app.ai.embeddings import embed_text as real_embed_text

def embed_text_sync(text: str) -> list[float]:
    for attempt in range(5):
        try:
            return real_embed_text(text, task_type="RETRIEVAL_DOCUMENT")
        except Exception as e:
            if "429" in str(e) or "quota" in str(e).lower():
                wait = 2 ** attempt * 5
                print(f"    Rate limited — waiting {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise Exception("Max retries exceeded for embedding")

async def embed_text(text: str) -> list[float]:
    return await asyncio.to_thread(embed_text_sync, text)

def parse_skillsets(wb) -> list[dict]:
    ws = wb["Skillset Reference"]
    chunks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == "Category":
            continue
        category_type, skill_name, aligned_sos, description = row
        chunks.append({
            "title": skill_name,
            "content": f"""BSCpE Key Skillset: {skill_name}.
Type: {category_type} Skill.
Aligned ABET Student Outcomes: {aligned_sos}.
Description: {description}
Career relevance: Students who score well in this skillset are strong candidates
for careers that require {skill_name.lower()} competency.
This skillset is assessed across multiple BSU CpE subjects through ILO scores."""
        })
    return chunks

def parse_student_outcomes(wb) -> list[dict]:
    ws = wb["SO Legend"]
    chunks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == "SO Code":
            continue
        so_code, so_name, full_definition = row
        chunks.append({
            "title": f"{so_code}: {so_name}",
            "content": f"""ABET Student Outcome {so_code}: {so_name}.
Full definition: {full_definition}
BSCpE context: This outcome is assessed through ILO scores across multiple
BSU CpE subjects. Students who consistently score high in {so_code}-aligned ILOs
demonstrate {so_name.lower()} competency.
Career implication: Strong {so_code} scores indicate readiness for careers
requiring {so_name.lower()} abilities in the Philippine tech industry."""
        })
    return chunks

def parse_curriculum(wb) -> list[dict]:
    ws = wb["ILO–Skillset Alignment"]
    chunks = []
    current_subject = None
    current_so = None
    current_ilos = []

    def flush_subject():
        if current_subject and current_ilos:
            ilo_text = "\n".join([
                f"  {i['ilo']}: {i['statement']}\n"
                f"    Technical skills: {i['tech']}\n"
                f"    Professional skills: {i['prof']}"
                for i in current_ilos
            ])
            chunks.append({
                "title": current_subject.strip(),
                "content": f"""BSU CpE Subject: {current_subject.strip()}.
Mapped Student Outcomes: {current_so or 'See ILO details'}.
Intended Learning Outcomes and Skills Developed:
{ilo_text}
Assessment context: Student scores in ILO1 through ILO{len(current_ilos)} of this subject
reflect their mastery of the technical and professional skills listed above.
High ILO scores in this subject predict competency in the mapped skill categories,
which in turn align with specific tech career paths."""
            })

    for row in ws.iter_rows(values_only=True):
        r0 = str(row[0]).strip() if row[0] else ''
        r2 = str(row[2]).strip() if row[2] else ''
        r3 = str(row[3]).strip() if row[3] else ''
        r4 = str(row[4]).strip() if row[4] else ''
        r5 = str(row[5]).strip() if row[5] else ''
        r6 = str(row[6]).strip() if row[6] else ''

        # Detect subject header
        is_subject = (
            row[1] is None and row[2] is None and row[0]
            and 'ILO' not in r0
            and 'BATANGAS' not in r0
            and 'Bachelor' not in r0
            and 'LEGEND' not in r0
            and 'Each ILO' not in r0
            and not r0.startswith('SO')
            and len(r0) > 3
            and '·' in r0
        )

        if is_subject:
            flush_subject()
            current_subject = r0
            current_so = None
            current_ilos = []

        elif r0.startswith('SO') and current_subject:
            current_so = r0.strip()

        elif r2.startswith('ILO') and current_subject and r3:
            current_ilos.append({
                'ilo': r2,
                'statement': r3,
                'tech': r5 if r5 and r5 != '—' else 'General engineering competency',
                'prof': r6 if r6 and r6 != '—' else 'Professional development'
            })

    flush_subject()
    return chunks

def parse_ml_meta() -> list[dict]:
    with open(ML_DIR / "meta.json") as f:
        meta = json.load(f)

    chunks = []
    for skill in meta.get("skills", meta.get("categories", meta.get("skill_categories", []))):
        skill_name = skill if isinstance(skill, str) else skill.get("name", str(skill))
        chunks.append({
            "title": f"ML Skill Category: {skill_name}",
            "content": f"""ASPIRE ML Model Skill Category: {skill_name}.
This skill category is one of the 20 categories predicted by ASPIRE's
GradientBoosting ML model based on a student's ILO scores.
The ML model predicts a proficiency score (0-100) for this category
by analyzing the student's ILO1, ILO2, ILO3, ILO4 scores across all
enrolled BSU CpE subjects.
Threshold classifications:
  Exceeding Expectations: score >= 80
  On Track: score >= 60
  Needs Attention: score >= 40
  Critical: score < 40
Career relevance: Strong scores in {skill_name} directly contribute to
career readiness scores for aligned tech career paths."""
        })
    return chunks

CAREER_PATHS = [
    {
        "title": "Backend Developer",
        "content": """Career Path: Backend Developer.
Also known as: Backend Engineer, Server-side Developer, API Developer, Backend Software Engineer.
Philippine market outlook: High demand — top tech career in PH 2024-2025.
Required skills: Python or JavaScript, REST APIs, PostgreSQL,
Docker, Git, Linux, Redis, CI/CD pipelines.
BSCpE skillsets aligned: Programming & Software Development,
Operating Systems & Architecture, Networking & Communications.
ABET SOs: SO1, SO5, SO11.
Roadmap: https://roadmap.sh/backend
Demonstrative GitHub projects: REST API with FastAPI/Express + PostgreSQL + JWT auth; webhook server with rate limiting; CRUD application with Docker Compose; URL shortener with Redis caching.
Learning order: Programming Language → Databases → REST APIs →
Authentication → Containerization → CI/CD → Cloud deployment.
BSU CpE subjects most relevant: CpE 406 OOP, CpE 411 Data Structures,
CpE 418 Software Design, CpE 424 Operating Systems, Database courses."""
    },
    {
        "title": "Full Stack Developer",
        "content": """Career Path: Full Stack Developer.
Also known as: Full Stack Engineer, Full Stack Software Developer, Web Developer, Full-stack Application Developer.
Philippine market outlook: Very high demand — most versatile tech role in PH.
Required skills: React, Node.js or Python, PostgreSQL, REST APIs,
Git, Docker, basic DevOps, HTML/CSS/JavaScript.
BSCpE skillsets aligned: Programming & Software Development,
Operating Systems & Architecture, Modern Engineering Tools.
ABET SOs: SO1, SO3, SO5, SO11.
Roadmap: https://roadmap.sh/full-stack
Demonstrative GitHub projects: SPA + REST API + database deployed end-to-end (React + Node/Python + Postgres); marketplace clone with auth; SaaS dashboard with Stripe-style billing; blog/CMS with admin panel.
Learning order: HTML/CSS/JS → React → Backend Language →
Database → APIs → Authentication → Deployment.
BSU CpE subjects most relevant: CpE 406 OOP, CpE 411 Data Structures,
CpE 418 Software Design, Web Systems and Technologies."""
    },
    {
        "title": "Frontend Developer",
        "content": """Career Path: Frontend Developer.
Also known as: Frontend Engineer, UI Developer, Web Frontend Developer, JavaScript Developer, React Developer.
Philippine market outlook: High demand especially for React developers.
Required skills: HTML, CSS, JavaScript, React or Vue,
TypeScript, Git, REST API integration, responsive design, accessibility.
BSCpE skillsets aligned: Programming & Software Development,
Modern Engineering Tools.
ABET SOs: SO1, SO5, SO11.
Roadmap: https://roadmap.sh/frontend
Demonstrative GitHub projects: Responsive React or Vue SPA with state management and routing; component library with Storybook; accessible form-heavy app (a11y compliant); progressive web app (PWA) with service workers; portfolio site with animations.
Learning order: HTML → CSS → JavaScript → React →
TypeScript → Build Tools → Testing → Performance.
BSU CpE subjects most relevant: CpE 401 Programming 1,
CpE 406 OOP, CpE 418 Software Design."""
    },
    {
        "title": "DevOps Engineer",
        "content": """Career Path: DevOps Engineer.
Also known as: Site Reliability Engineer (SRE), Platform Engineer, Cloud Engineer, Infrastructure Engineer, Build & Release Engineer.
Philippine market outlook: Growing — critical shortage of qualified DevOps in PH.
Required skills: Linux, Docker, Kubernetes, CI/CD,
AWS or GCP, Terraform, monitoring tools, bash scripting, networking.
BSCpE skillsets aligned: Operating Systems & Architecture,
Networking & Communications, Modern Engineering Tools.
ABET SOs: SO1, SO3, SO5, SO11.
Roadmap: https://roadmap.sh/devops
Demonstrative GitHub projects: Multi-container docker-compose stack; CI/CD pipeline with GitHub Actions or Jenkins; Terraform module deploying AWS/GCP infrastructure; Kubernetes manifest set with Helm chart; Prometheus/Grafana monitoring setup.
Learning order: Linux → Networking → Git → Docker →
CI/CD → Kubernetes → Cloud → Monitoring → Infrastructure as Code.
BSU CpE subjects most relevant: CpE 412 CISCO 1, CpE 419 CISCO 2,
CpE 423 CISCO 3, CpE 424 Operating Systems, CpE 427 CISCO 4."""
    },
    {
        "title": "Cybersecurity Analyst",
        "content": """Career Path: Cybersecurity Analyst.
Also known as: Security Analyst, SOC Analyst, Information Security Analyst, Penetration Tester, Cybersecurity Engineer.
Philippine market outlook: Critical demand — PH ranks high in cyberattack targets.
Required skills: Networking, Linux, ethical hacking, penetration testing,
SIEM tools, incident response, security frameworks, cryptography.
BSCpE skillsets aligned: Networking & Communications,
Operating Systems & Architecture, Ethics & Professionalism.
ABET SOs: SO1, SO2, SO3, SO5, SO6.
Roadmap: https://roadmap.sh/cyber-security
Demonstrative GitHub projects: CTF challenge writeups; vulnerability scanner script; SIEM/log analysis dashboard with ELK stack; small-scale penetration testing report on a vulnerable VM (e.g., Metasploitable); home network IDS with Suricata.
Learning order: Networking Fundamentals → Linux → Security Basics →
Ethical Hacking → Penetration Testing → Incident Response → Certifications.
BSU CpE subjects most relevant: CpE 412-427 CISCO track,
CpE 424 Operating Systems, CpE 427 Connecting Networks and Security."""
    },
    {
        "title": "Data Scientist",
        "content": """Career Path: Data Scientist.
Also known as: Data Analyst, Quantitative Analyst, Research Scientist, Applied Scientist, Analytics Engineer.
Philippine market outlook: High growth — banking, BPO, and tech sectors.
Required skills: Python, pandas, scikit-learn, SQL,
data visualization, statistical analysis, machine learning, Jupyter.
BSCpE skillsets aligned: Data Science & AI/ML,
Mathematics & Science Foundations, Programming & Software Development.
ABET SOs: SO1, SO2, SO5.
Roadmap: https://roadmap.sh/ai-data-scientist
Demonstrative GitHub projects: Jupyter notebook with EDA + ML model + visualization (Kaggle-style competition); time-series forecasting on real PH datasets (PSE stocks, weather); A/B test analysis report; recommendation system; SQL analytics dashboard.
Learning order: Python → Statistics → SQL → Data Analysis →
Machine Learning → Deep Learning → Model Deployment.
BSU CpE subjects most relevant: MATH 401-404, ENGG 414 Numerical Methods,
CpEE 403 Data Mining/AI Elective, MATH 403 Engineering Data Analysis."""
    },
    {
        "title": "AI Engineer",
        "content": """Career Path: AI Engineer.
Also known as: GenAI Engineer, LLM Engineer, ML Platform Engineer, Applied AI Engineer, AI Software Engineer.
Philippine market outlook: Emerging — fastest growing tech role globally.
Required skills: Python, PyTorch or TensorFlow, LLMs,
prompt engineering, vector databases, RAG pipelines, MLOps.
BSCpE skillsets aligned: Data Science & AI/ML,
Mathematics & Science Foundations, Programming & Software Development.
ABET SOs: SO1, SO2, SO3, SO5.
Roadmap: https://roadmap.sh/ai-engineer
Demonstrative GitHub projects: RAG chatbot with vector database (pgvector/Pinecone); LLM fine-tuning pipeline; multi-agent framework using LangChain/CrewAI; prompt-engineering experiments; semantic search engine; RAG-based Q&A on custom documents.
Learning order: Python → ML Fundamentals → Deep Learning →
Transformers/LLMs → Prompt Engineering → RAG → MLOps → Cloud AI.
BSU CpE subjects most relevant: CpEE 403 Data Mining/AI Elective,
ENGG 414 Numerical Methods, MATH 403 Engineering Data Analysis."""
    },
    {
        "title": "Machine Learning Engineer",
        "content": """Career Path: Machine Learning Engineer.
Also known as: ML Engineer, Applied Machine Learning Engineer, MLOps Engineer, Production ML Engineer.
Philippine market outlook: Growing — fintech and healthtech sectors leading.
Required skills: Python, scikit-learn, TensorFlow or PyTorch,
feature engineering, model deployment, Docker, REST APIs for ML, MLflow.
BSCpE skillsets aligned: Data Science & AI/ML,
Mathematics & Science Foundations, Engineering Design & Research.
ABET SOs: SO1, SO2, SO3, SO5.
Roadmap: https://roadmap.sh/machine-learning
Demonstrative GitHub projects: scikit-learn pipeline deployed via FastAPI/Flask; MLflow experiment tracking with reproducible runs; feature store demo; automated retraining pipeline; model monitoring dashboard with drift detection.
Learning order: Python → Statistics → ML Algorithms →
Model Training → Feature Engineering → Model Deployment → MLOps.
BSU CpE subjects most relevant: CpEE 403 Data Mining/AI Elective,
ENGG 414 Numerical Methods, CpE 411 Data Structures."""
    },
    {
        "title": "Software Architect",
        "content": """Career Path: Software Architect.
Also known as: Principal Engineer, Tech Lead, Solutions Architect, Senior Software Engineer (Architecture track), Staff Engineer.
Philippine market outlook: Senior role — 5+ years experience typically required.
Required skills: System design, microservices, distributed systems,
cloud architecture, API design patterns, database design,
security architecture, performance optimization.
BSCpE skillsets aligned: Programming & Software Development,
Engineering Design & Research, Operating Systems & Architecture.
ABET SOs: SO1, SO3, SO5, SO11, SO12.
Roadmap: https://roadmap.sh/software-architect
Demonstrative GitHub projects: System design documents (architecture diagrams + ADRs); reference microservices implementation with API gateway; performance optimization case study with benchmarks; distributed system demo (e.g., consensus algorithm, leader election).
Learning order: Strong backend foundation → System Design →
Distributed Systems → Cloud → Security → Leadership.
BSU CpE subjects most relevant: CpE 418 Software Design,
CpE 411 Data Structures, CpE 422 CpE Practice and Design 1,
CpE 430 CpE Practice and Design 2, Capstone."""
    }
]

HARDWARE_CAREER_PATHS = [
    {
        "title": "Embedded Systems Engineer",
        "content": """Career Path: Embedded Systems Engineer.
Also known as: Firmware Engineer, Microcontroller Engineer, Hardware-Software Integration Engineer, Embedded Software Engineer.
Philippine market outlook: Steady demand in manufacturing and consumer electronics.
Required skills: C/C++, RTOS (FreeRTOS/Zephyr), Microcontrollers (STM32, ESP32, AVR),
I2C/SPI/UART protocols, PCB design basics, debugging with Oscilloscopes.
BSCpE skillsets aligned: Embedded & Microprocessor Systems,
Hardware & Circuit Design, Operating Systems & Architecture.
ABET SOs: SO1, SO2, SO3, SO5.
Roadmap: https://roadmap.sh/embedded
Demonstrative GitHub projects: STM32/ESP32 firmware with peripherals (I2C/SPI sensor reading); RTOS-based scheduler demo (FreeRTOS tasks); bootloader implementation; sensor fusion project (IMU + Kalman filter); custom HAL driver.
Learning order: C programming → Digital Logic → Microcontroller fundamentals → Communication Protocols (I2C/SPI/UART) → RTOS → Embedded Linux → IoT Integration.
BSU CpE subjects most relevant: CpE 417 Microprocessors,
CpE 428 Embedded Systems, CpE 413 Fundamentals of Mixed Signals and Sensors."""
    },
    {
        "title": "VLSI Design Engineer",
        "content": """Career Path: VLSI Design Engineer.
Also known as: Digital Design Engineer, ASIC Design Engineer, Chip Design Engineer, RTL Design Engineer.
Philippine market outlook: Specialized demand in semiconductor design houses (Laguna/Cavite).
Required skills: Verilog/SystemVerilog, VHDL, Digital Logic Design,
CMOS fundamentals, EDA tools (Cadence, Synopsys), Static Timing Analysis.
BSCpE skillsets aligned: Hardware & Circuit Design,
Embedded & Microprocessor Systems, Engineering Design & Research.
ABET SOs: SO1, SO2, SO3.
Roadmap: https://roadmap.sh/computer-science
Demonstrative GitHub projects: Verilog modules (ALU, FIFO, AXI bus master) with testbenches; UVM verification environment; pipelined RISC-V core implementation; synthesis and STA reports.
Learning order: Digital Logic → HDL fundamentals (Verilog/VHDL) → CMOS basics → Synthesis → EDA tooling → Static Timing Analysis → Verification (UVM) → Physical Design.
BSU CpE subjects most relevant: CpE 410 Logic Circuits and Design,
CpE 415 Introduction to HDL, CpE 425 Computer Architecture and Organization."""
    },
    {
        "title": "FPGA Engineer",
        "content": """Career Path: FPGA Engineer.
Also known as: FPGA Developer, Hardware Engineer (FPGA), Digital Hardware Engineer, RTL Engineer, FPGA Firmware Engineer.
Philippine market outlook: Growing niche in telecommunications and high-speed trading.
Required skills: HDL (Verilog/VHDL), FPGA Architecture (Xilinx/Altera),
Digital Signal Processing, High-speed interfaces, Timing closure.
BSCpE skillsets aligned: Hardware & Circuit Design,
Signal Processing & Control Systems, Modern Engineering Tools.
ABET SOs: SO1, SO2, SO3, SO5.
Roadmap: https://roadmap.sh/computer-science
Demonstrative GitHub projects: Xilinx Vivado/Altera Quartus designs; DSP module on FPGA (FIR/IIR filters); high-speed serial protocol implementation (UART, SPI, AXI); FPGA-based image processing demo; HDMI/VGA controller.
Learning order: Digital Logic → HDL (Verilog/VHDL) → Synthesis → FPGA tooling (Vivado/Quartus) → Timing closure → DSP fundamentals → High-speed interfaces.
BSU CpE subjects most relevant: CpE 415 Introduction to HDL,
CpE 420 Digital Signal Processing, CpE 410 Logic Circuits and Design."""
    },
    {
        "title": "IoT Systems Engineer",
        "content": """Career Path: IoT Systems Engineer.
Also known as: IoT Developer, Connected Devices Engineer, Edge Computing Engineer, Smart Systems Engineer, IoT Solutions Architect.
Philippine market outlook: High potential in smart agriculture and industrial automation.
Required skills: Embedded C, Wireless protocols (LoRaWAN, Zigbee, BLE),
Cloud integration (AWS IoT, Google Cloud IoT), Sensors, MQTT.
BSCpE skillsets aligned: Networking & Communications,
Embedded & Microprocessor Systems, Data Science & AI/ML.
ABET SOs: SO1, SO3, SO5, SO11.
Roadmap: https://roadmap.sh/embedded
Demonstrative GitHub projects: ESP32/Arduino with MQTT to AWS IoT or Google Cloud IoT; sensor data dashboard (Grafana + InfluxDB); LoRaWAN gateway demo; smart home prototype with multiple devices; edge ML on microcontroller (TensorFlow Lite Micro).
Learning order: Embedded fundamentals → Wireless protocols → Sensors and actuators → Cloud integration → MQTT/CoAP → Edge processing → Security for IoT.
BSU CpE subjects most relevant: CpE 412-427 CISCO track,
CpE 428 Embedded Systems, CpE 413 Fundamentals of Mixed Signals and Sensors."""
    }
]

GAP_CLOSER_PROJECTS = [
    {
        "title": "Dockerization Project for BSCpE",
        "content": """Project: "Containerized Microservice Deployment"
Target Gap: Docker & Containerization.
Description: Take an existing Python (FastAPI) or Node.js project and
create a Dockerfile. Use docker-compose to link it with a PostgreSQL database.
Key learning: Images, Containers, Volumes, Networking, Docker Compose.
Relevance: Critical for Backend, DevOps, and Full Stack roles."""
    },
    {
        "title": "CI/CD Pipeline Project for BSCpE",
        "content": """Project: "Automated Testing and Linting Pipeline"
Target Gap: CI/CD.
Description: Create a GitHub repository and set up GitHub Actions.
Configure it to run automated tests (Pytest/Jest) and linting (Flake8/ESLint)
on every pull request.
Key learning: YAML configuration, GitHub Actions, Workflow automation.
Relevance: Essential for modern software engineering and DevOps."""
    },
    {
        "title": "Cloud Deployment Project for BSCpE",
        "content": """Project: "Full-stack Production Deployment"
Target Gap: Cloud Infrastructure.
Description: Deploy a frontend (React) to Vercel and a backend (FastAPI)
to Render or AWS App Runner. Configure environment variables and CORS.
Key learning: DNS, SSL/TLS, Environment management, Cloud providers.
Relevance: Necessary for all web-focused career paths."""
    }
]

PH_TECH_ECOSYSTEM = [
    {
        "title": "Philippine Tech Hubs",
        "content": """Knowledge: Major Tech Hubs in the Philippines.
Locations: Makati City (Fintech/Banking), BGC Taguig (Multinationals/Startups),
Ortigas Center (BPO/IT Services), Cebu City (Emerging Tech/Software Houses).
Significance: Understanding where companies are located helps in job searching
and deciding on relocation or remote work opportunities."""
    },
    {
        "title": "Philippine Tech Communities",
        "content": """Knowledge: Active Tech Communities for Networking in PH.
Groups: Python Philippines, Google Developer Groups (GDG) Philippines,
AWS User Group Philippines, React Philippines, Women Who Code Manila.
Advice: Joining these groups on Facebook, Meetup, or Discord helps students
find mentors and stay updated on local industry trends."""
    }
]

PROFESSIONAL_SKILLS_MAPPING = [
    {
        "title": "Professional Communication for Engineers",
        "content": """Competency: Technical Communication.
Aligned Subject: GEd 106 Purposive Communication.
Description: Beyond basic speaking, this subject prepares BSCpE students
for technical writing, client presentations, and documentation.
Relevance: Essential for Senior roles, Project Management, and Client-facing engineering."""
    },
    {
        "title": "Ethics in the Digital Age",
        "content": """Competency: Ethics & Professionalism.
Aligned Subject: GEd 107 Ethics.
Description: Focuses on engineering ethics, data privacy (Data Privacy Act of 2012),
and professional responsibility in software and hardware design.
Relevance: Critical for Cybersecurity, AI Engineering, and Lead Developer roles."""
    }
]

LEARNING_RESOURCES = [
    {
        "title": "Free Learning Resources for BSCpE Gaps",
        "content": """Resources: Recommended Platforms for Self-Paced Learning.
- Roadmap.sh: Comprehensive visual roadmaps for all tech roles.
- FreeCodeCamp: Project-based learning for Web Development and Data Science.
- MDN Web Docs: The gold standard for frontend (HTML/CSS/JS) documentation.
- Cisco Networking Academy: Deep dives into networking and security."""
    }
]


ROADMAP_LEARNING_PATHS = [
    {
        "title": "Backend Developer learning path for BSU CpE",
        "content": """Backend Developer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Already covered in BSCpE curriculum:
  Internet basics HTTP DNS, Linux fundamentals (CpE 424 Operating Systems),
  Git version control, Networking fundamentals (CISCO 1-4 track).
Stage 2 — Programming proficiency (CpE 401, CpE 406 OOP):
  Python or JavaScript as primary backend language.
  OOP patterns, data structures (CpE 411), algorithms.
Stage 3 — Databases (Database Management subject):
  PostgreSQL — relational design, SQL, indexing, transactions.
  ORM — SQLAlchemy. Redis for caching.
Stage 4 — APIs:
  REST API design and implementation using FastAPI or Express.
  Authentication — JWT, OAuth 2.0.
Stage 5 — Containerization (gap for most BSU CpE students):
  Docker — containers, images, docker-compose.
Stage 6 — CI/CD (gap for most BSU CpE students):
  GitHub Actions for automated testing and deployment.
Stage 7 — Cloud (gap for most BSU CpE students):
  AWS or GCP fundamentals, deploy a production backend."""
    },
    {
        "title": "DevOps Engineer learning path for BSU CpE",
        "content": """DevOps Engineer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Linux system administration (CpE 424 Operating Systems).
  Networking — TCP/IP, DNS, routing (CISCO 1-4 track CpE 412-427).
  Shell scripting — bash.
Stage 2 — Version control:
  Git workflows, branching strategies, GitHub or GitLab.
Stage 3 — Containerization (gap skill):
  Docker — images, containers, docker-compose, container security.
Stage 4 — CI/CD (gap skill):
  GitHub Actions, Jenkins or GitLab CI — automated pipelines.
Stage 5 — Cloud (gap skill):
  AWS or GCP fundamentals — EC2, S3, VPC, IAM.
Stage 6 — Orchestration (advanced gap):
  Kubernetes — pods, deployments, services, Helm charts.
Stage 7 — Monitoring (advanced gap):
  Prometheus, Grafana, logging and alerting systems."""
    },
    {
        "title": "Data Scientist learning path for BSU CpE",
        "content": """Data Scientist roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Mathematics — calculus (MATH 401-402), statistics (MATH 403),
  differential equations (MATH 404), numerical methods (ENGG 414).
Stage 2 — Python for data:
  pandas, numpy for data manipulation.
  matplotlib, seaborn for visualization.
Stage 3 — Machine learning (partially covered in CpEE 403):
  scikit-learn, model training and evaluation.
  Regression, classification, clustering algorithms.
Stage 4 — SQL and databases:
  Data querying, aggregation, working with large datasets.
Stage 5 — Advanced ML (gap skill):
  Neural networks, TensorFlow or PyTorch.
  Feature engineering, cross-validation, hyperparameter tuning.
Stage 6 — Deployment (gap skill):
  Model serving via REST APIs, MLflow for experiment tracking."""
    },
    {
        "title": "AI Engineer learning path for BSU CpE",
        "content": """AI Engineer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Python, mathematics, engineering data analysis (MATH 403).
  Numerical methods (ENGG 414), data science elective (CpEE 403).
Stage 2 — Machine learning fundamentals:
  scikit-learn, model training, evaluation metrics.
Stage 3 — Deep learning (gap skill):
  Neural networks, backpropagation, PyTorch or TensorFlow.
Stage 4 — LLMs and generative AI (gap skill):
  Transformer architecture, prompt engineering, LangChain.
Stage 5 — RAG and vector databases (gap skill):
  Embeddings, semantic search, pgvector, ChromaDB.
Stage 6 — MLOps and deployment (gap skill):
  Model serving, monitoring, API wrapping for AI models.
Stage 7 — Cloud AI (gap skill):
  Google Vertex AI, AWS SageMaker, cost optimization."""
    },
    {
        "title": "Cybersecurity Analyst learning path for BSU CpE",
        "content": """Cybersecurity Analyst roadmap — ordered stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Networking (CISCO 1-4 track — CpE 412, 419, 423, 427).
  Operating Systems (CpE 424), Ethics (GEd 107).
Stage 2 — Security fundamentals:
  CIA triad, common attack vectors, cryptography basics.
Stage 3 — Linux security (partially covered):
  Hardening, permissions, file system security.
Stage 4 — Ethical hacking (gap skill):
  Penetration testing, Kali Linux, vulnerability scanning.
Stage 5 — Incident response (gap skill):
  SIEM tools, log analysis, forensics basics.
Stage 6 — Certifications (gap skill):
  CompTIA Security+, CEH, or CCNA Security."""
    },
    {
        "title": "Full Stack Developer learning path for BSU CpE",
        "content": """Full Stack Developer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — BSCpE foundation:
  Data structures and algorithms (CpE 411).
  OOP principles (CpE 406 OOP).
  Software design (CpE 418).
Stage 2 — Frontend basics (gap skill):
  HTML, CSS, JavaScript fundamentals.
  Responsive design, DOM manipulation.
Stage 3 — Frontend frameworks (gap skill):
  React or Vue.js. State management, routing.
Stage 4 — Backend APIs (partially covered):
  Node.js or Python (FastAPI/Django).
  RESTful API design, authentication.
Stage 5 — Databases (Database Management subject):
  Relational databases (PostgreSQL/MySQL), SQL queries.
Stage 6 — Full Stack Integration (gap skill):
  Connecting frontend to backend, handling CORS, deployment strategies (Vercel, Heroku, AWS)."""
    },
    {
        "title": "Frontend Developer learning path for BSU CpE",
        "content": """Frontend Developer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — BSCpE foundation:
  Basic programming logic (CpE 401, CpE 404).
  Software design principles (CpE 418).
Stage 2 — Core web technologies (gap skill):
  Semantic HTML5, CSS3, modern JavaScript (ES6+).
Stage 3 — UI/UX and styling (gap skill):
  CSS frameworks (TailwindCSS, Bootstrap), responsive design.
Stage 4 — Component-driven development (gap skill):
  React, Vue, or Angular. Managing props, state, and lifecycle.
Stage 5 — Advanced state and routing (gap skill):
  Redux, Context API, React Router.
Stage 6 — Web performance and testing (gap skill):
  Lighthouse optimization, Jest, Cypress testing.
Stage 7 — Progressive Web Apps (PWA) (gap skill):
  Service workers, offline support."""
    },
    {
        "title": "Software Architect learning path for BSU CpE",
        "content": """Software Architect roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Data structures (CpE 411), Object-oriented programming (CpE 406).
  Software design (CpE 418), Operating Systems (CpE 424).
Stage 2 — Backend mastery (advanced):
  Deep expertise in backend languages, databases, and APIs.
Stage 3 — System design fundamentals (gap skill):
  Scalability, availability, CAP theorem, load balancing.
Stage 4 — Architecture patterns (gap skill):
  Microservices vs Monoliths, event-driven architecture.
  Message queues (RabbitMQ, Kafka).
Stage 5 — Cloud and infrastructure (gap skill):
  AWS/GCP/Azure architecture, container orchestration (Kubernetes).
Stage 6 — Security and compliance (partially covered in CISCO 4):
  Zero-trust architecture, OAuth 2.0, data encryption.
Stage 7 — Technical leadership (developed in Capstone):
  Mentoring, technology selection, resolving technical debt."""
    },
    {
        "title": "Machine Learning Engineer learning path for BSU CpE",
        "content": """Machine Learning Engineer roadmap — ordered learning stages for BSU CpE students.
Stage 1 — Strong BSCpE foundation:
  Python programming, Engineering data analysis (MATH 403), Numerical methods (ENGG 414).
Stage 2 — Data manipulation basics:
  pandas, numpy, data cleaning and preprocessing.
Stage 3 — Core machine learning (CpEE 403 Data Mining/AI Elective):
  scikit-learn, regression, classification, clustering.
Stage 4 — Advanced ML and Deep Learning (gap skill):
  Neural networks, CNNs, RNNs, PyTorch or TensorFlow.
Stage 5 — MLOps fundamentals (gap skill):
  MLflow, model versioning, experiment tracking.
Stage 6 — API engineering (partially covered):
  Deploying models behind REST APIs (FastAPI/Flask).
Stage 7 — Cloud ML deployment (gap skill):
  Dockerizing ML models, deploying on AWS SageMaker or GCP Vertex AI."""
    }
]

async def seed_category(
    session: AsyncSession,
    category: str,
    chunks: list[dict]
) -> int:
    print(f"\n  Seeding [{category}] — {len(chunks)} chunks...")
    count = 0
    for i, chunk in enumerate(chunks):
        try:
            embedding = await embed_text(chunk["content"])
            knowledge = KnowledgeChunk(
                category=category,
                title=chunk["title"],
                content=chunk["content"],
                embedding=embedding
            )
            session.add(knowledge)
            count += 1
            if (i + 1) % 5 == 0:
                await session.commit()
                print(f"    {i+1}/{len(chunks)} embedded...")
            time.sleep(0.5)  # respect Gemini rate limits
        except Exception as e:
            print(f"    ERROR on chunk '{chunk['title']}': {e}")
            continue
    await session.commit()
    print(f"  ✅ [{category}] — {count}/{len(chunks)} seeded")
    return count

async def main():
    print("ASPIRE Knowledge Base Seeder")
    print("=" * 40)

    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("\nParsing documents...")
    skillsets = parse_skillsets(wb)
    outcomes = parse_student_outcomes(wb)
    curriculum = parse_curriculum(wb)
    ml_skills = parse_ml_meta()

    total = (
        len(skillsets) + len(outcomes) + len(curriculum) +
        len(ml_skills) + len(CAREER_PATHS) + len(ROADMAP_LEARNING_PATHS) +
        len(HARDWARE_CAREER_PATHS) + len(GAP_CLOSER_PROJECTS) +
        len(PH_TECH_ECOSYSTEM) + len(PROFESSIONAL_SKILLS_MAPPING) +
        len(LEARNING_RESOURCES)
    )
    print(f"Total chunks to seed: {total}")
    print(f"  skillset: {len(skillsets)}")
    print(f"  ilo: {len(outcomes)}")
    print(f"  curriculum: {len(curriculum)}")
    print(f"  skillset_ml: {len(ml_skills)}")
    print(f"  career_path: {len(CAREER_PATHS) + len(HARDWARE_CAREER_PATHS)}")
    print(f"  roadmap: {len(ROADMAP_LEARNING_PATHS)}")
    print(f"  gap_closer: {len(GAP_CLOSER_PROJECTS)}")
    print(f"  ph_tech_eco: {len(PH_TECH_ECOSYSTEM)}")
    print(f"  prof_skills: {len(PROFESSIONAL_SKILLS_MAPPING)}")
    print(f"  resources: {len(LEARNING_RESOURCES)}")

    async with async_session_factory() as session:
        print("\nClearing existing knowledge chunks...")
        await session.execute(delete(KnowledgeChunk))
        await session.commit()
        print("  Cleared")

        total_seeded = 0
        total_seeded += await seed_category(session, "skillset", skillsets)
        total_seeded += await seed_category(session, "ilo", outcomes)
        total_seeded += await seed_category(session, "curriculum", curriculum)
        total_seeded += await seed_category(session, "skillset_ml", ml_skills)
        total_seeded += await seed_category(session, "career_path", CAREER_PATHS)
        total_seeded += await seed_category(session, "career_path", HARDWARE_CAREER_PATHS)
        total_seeded += await seed_category(session, "roadmap", ROADMAP_LEARNING_PATHS)
        total_seeded += await seed_category(session, "gap_closer", GAP_CLOSER_PROJECTS)
        total_seeded += await seed_category(session, "ph_tech_eco", PH_TECH_ECOSYSTEM)
        total_seeded += await seed_category(session, "prof_skills", PROFESSIONAL_SKILLS_MAPPING)
        total_seeded += await seed_category(session, "resources", LEARNING_RESOURCES)

    print("\n" + "=" * 40)
    print(f"Knowledge base seeded: {total_seeded}/{total} chunks")
    print("Run 'python scripts/seed_knowledge.py' again to re-seed")

if __name__ == "__main__":
    asyncio.run(main())
